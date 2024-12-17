import openpyxl
from adjacency_list_graph import AdjacencyListGraph
from dijkstra import dijkstra
from print_path import print_path
import re
import matplotlib.pyplot as plt

def normalize_name(name):
    '''Remove extra spaces and convert to lowercase'''
    name = re.sub(r'\s+', ' ', name.strip().lower())
    return name

# Load London Underground excel sheet
# Uses openpyxl to load the excel sheet in our code
# It can happen an error 
wb = openpyxl.load_workbook('database.xlsx')
# Selects the first sheet from the excel 
sheet = wb.active

# Defining the structures needed as sets to avoid repeated values
lines_set = set()
stations = set()
connections = set()

# iter_rows is a method from openpyxl that allows us to select only the data needed from the excel
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=4):
    # Adds the first cell value to the set - column A - line names
    line = row[0].value
    lines_set.add(line)

    station1 = row[1].value  # Column B (first station)
    station2 = row[2].value  # Column C (second station)
    distance = row[3].value  # Column D (distance)

    # if column C is empty and column B is not, add the value of column B to the stations set 
    if station2 is None and station1 is not None:
        normalized_station = normalize_name(station1) # we need to normalize the data because some of the inputs have spaces or different capitalization
        stations.add(normalized_station)

    # if column C is not empty then add a tuple to the connections set
    if distance is not None:
        normalized_station1 = normalize_name(station1)
        normalized_station2 = normalize_name(station2) 
        
        # creates a sorted tuple of stations to ensure unique pairs (A,B and B,A are treated as the same)
        pair = tuple(sorted((normalized_station1, normalized_station2)))
        
        # checks for an existing connection to avoid duplicates, keeping only the one with the shortest distance -> (A,B,d1) where d1 > d0
        existing_connection = next((c for c in connections if c[0] == pair[0] and c[1] == pair[1]), None)
        if existing_connection:
            if distance < existing_connection[2]: 
                connections.remove(existing_connection) # removes the greater distance
                connections.add((pair[0], pair[1], distance)) # adds the smallest distance
        else:
            connections.add((pair[0], pair[1], distance)) # if there is no repeated distances just add it

# create a graph and add the edges
connections = list(connections)
stations = list(stations)

# generate a graph with vertices being the stations on the excel
graph = AdjacencyListGraph(len(stations), False, True)
for edge in connections:
    station1_index = stations.index(edge[0])
    station2_index = stations.index(edge[1])

    if not graph.has_edge(station1_index, station2_index):
        graph.insert_edge(station1_index, station2_index, edge[2])

# track the longest journey
longest_distance = -1  
longest_path = None
furthest_station_pair = None

# initialize counter for journeys
journey_counter = 0

# collection of all distances for histogram
all_distances = []

# run dijkstra for each station as the source
for s1_index in range(len(stations)):
    # run dijkstra for the current source station
    d, pi = dijkstra(graph, s1_index)
    # for each station, find the valid distances (excluding self loops and inf distances)
    for s2_index in range(s1_index + 1, len(stations)):  # ensure we only calculate for s1_index < s2_index
        # update journey counter
        journey_counter += 1
        # finding the longest distance
        if s1_index != s2_index:
            if d[s2_index] > longest_distance:
                longest_distance = d[s2_index]
                furthest_station_pair = (stations[s1_index], stations[s2_index])
                
                # Get the path for the longest journey
                path = print_path(pi, s1_index, s2_index, lambda v: stations[v])
                longest_path = path
        #ensure distances are valid (excluding infinite distances representing unreachable stations)
        # we add these distances to 'all_distances' for plotting the histogram later
        if s1_index != s2_index and d[s2_index] < float('inf'):
            all_distances.append(d[s2_index])

# print the number of journeys calculated
print(f"Total number of journeys calculated: {journey_counter}")

# print the longest journey duration and its path
if furthest_station_pair:
    print(f"Longest journey duration: {longest_distance} minutes")
    print(f"Path from {furthest_station_pair[0]} to {furthest_station_pair[1]}: {' -> '.join(longest_path)}")
else:
    print("No journey durations found.")


# Histogram:
plt.figure(figsize=(10, 6))
plt.hist(all_distances, bins=50, color='skyblue', edgecolor='black')  # More bins for finer granularity
plt.title("Histogram of Distances Between Station Pairs")
plt.xlabel("Distance (minutes)")
plt.ylabel("Number of Journeys")
plt.show()
