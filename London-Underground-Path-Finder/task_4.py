import openpyxl
from adjacency_list_graph import AdjacencyListGraph
from mst import kruskal  # Import Kruskal's algorithm from mst module
import re
from dijkstra import dijkstra
from print_path import print_path
import matplotlib.pyplot as plt

def normalize_name(name):
    '''Remove extra spaces and convert to lowercase'''
    name = re.sub(r'\s+', ' ', name.strip().lower())
    return name

# load London Underground excel sheet
excel = r'C:\Users\ainho\OneDrive\Escritorio\AADD Project\AA&DE\database.xlsx'
# uses openpyxl to load the excel sheet in our code
wb = openpyxl.load_workbook(excel)
# selects the first sheet from the excel 
sheet = wb.active

# Defining the structures needed as sets to avoid repeated values
lines_set = set()
stations = set()
connections = set()

# iter_rows is a method from openpyxl that allows us to select only the data needed from the excel
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=4):
    # adds the first cell value to the set - column A - line names
    line = row[0].value
    lines_set.add(line)

    station1 = row[1].value  # Column B (first station)
    station2 = row[2].value  # Column C (second station)
    distance = row[3].value  # Column D (distance)

    # if column C is empty and column B is not, add the value of column B to the stations set 
    if station2 is None and station1 is not None:
        normalized_station = normalize_name(station1) # we need to normalize the data because some of the inputs have spaces or different capitalization
        stations.add(normalized_station)

    # if column C is not empty then add a tuple to the connections set
    if distance is not None:
        normalized_station1 = normalize_name(station1)
        normalized_station2 = normalize_name(station2) 
        
        # creates a sorted tuple of stations to ensure unique pairs (A,B and B,A are treated as the same)
        pair = tuple(sorted((normalized_station1, normalized_station2)))
        
        # checks for an existing connection to avoid duplicates, keeping only the one with the shortest distance -> (A,B,d1) where d1 > d0
        existing_connection = next((c for c in connections if c[0] == pair[0] and c[1] == pair[1]), None)
        if existing_connection:
            if distance < existing_connection[2]: 
                connections.remove(existing_connection) # removes the greater distance
                connections.add((pair[0], pair[1], distance)) # adds the smallest distance
        else:
            connections.add((pair[0], pair[1], distance)) # if there is no repeated distances just add it

# convert to lists for indexing
connections = list(connections)
stations = list(stations)


stations.sort()  # ensure consistent station ordering - consistency in the output

# generate a graph with vertices being the stations on the excel
graph = AdjacencyListGraph(len(stations), False, True)
# insert the edges - connections between stations using the provided method from adjacency_list_graph
for edge in connections:
    station1_index = stations.index(edge[0])
    station2_index = stations.index(edge[1])

    # checks if the edge already exists - ensures no repetition
    if not graph.has_edge(station1_index, station2_index):
        graph.insert_edge(station1_index, station2_index, edge[2])

# calculate the Minimum Spanning Tree using Kruskal's algorithm, it will obtain 
# the essential connections (the ones that we cannot close in order to maintain connectivity)
new_graph = kruskal(graph)  

# set to store the essential connections (edges)
mst_edges = set() 
# loop through each vertex in the graph
for u in range(new_graph.get_card_V()):
    # for each station 'u', iterate through its adjacent vertices (connections)
    for edge in new_graph.get_adj_list(u):
        # get the adjacent station 'v' connected to station 'u'
        v = edge.get_v()
        # store each MST edge as a sorted tuple for comparison
        mst_edges.add(tuple(sorted((stations[u], stations[v]))))

# array to store the closed connections - tuples
closed_connections = []  
# for each edge in connections compare it to the mst_edges 
# and if it is not present then it means it is not essential and can be closed (add to closed connections array)
for edge in connections:
    edge_tuple = tuple(sorted((edge[0], edge[1])))
    if edge_tuple not in mst_edges:
        closed_connections.append(edge_tuple)

# in order to update the graph with the closed stations
# we will remove from the graph every edge existing in closed_connections
# but first we will check its existence in order to avoid errors
for edge in closed_connections:
    station1_index = stations.index(edge[0])
    station2_index = stations.index(edge[1])
    
    if graph.has_edge(station1_index, station2_index):
        graph.delete_edge(station1_index, station2_index)

# Now `graph` only contains essential edges that are part of the MST
# Closed stations
print("Closed Stations:")
for connection in closed_connections:
    print(f"{connection[0].title()} -- {connection[1].title()}")

# Tracking the longest journey
longest_distance = -1  
longest_path = None
furthest_station_pair = None
all_distances = []  # collection of all distances for histogram

# run dijkstra for each station as the source
for s1_index in range(len(stations)):
    # run dijkstra for the current source station
    d, pi = dijkstra(graph, s1_index)
    # for each station, find the valid distances (excluding self loops and inf distances)
    for s2_index in range(s1_index + 1, len(stations)):  # this ensures we only calculate for s1_index < s2_index | excludind repeated undirected edges A - B and B - A 
        if s1_index != s2_index:
            # finding the longest distance
            if d[s2_index] > longest_distance:
                longest_distance = d[s2_index]
                furthest_station_pair = (stations[s1_index], stations[s2_index])
                
                # get the path for the longest journey
                path = print_path(pi, s1_index, s2_index, lambda v: stations[v])
                longest_path = path
        #ensure distances are valid (excluding infinite distances representing unreachable stations)
        # we add these distances to 'all_distances' for plotting the histogram later
        if s1_index != s2_index and d[s2_index] < float('inf'):
            all_distances.append(d[s2_index])

# print the longest journey duration and its path
if furthest_station_pair:
    print(f"Longest journey duration: {longest_distance} minutes")
    print(f"Path from {furthest_station_pair[0]} to {furthest_station_pair[1]}: {' -> '.join(longest_path)}")
else:
    print("No journey durations found.")

# Histogram:
plt.figure(figsize=(10, 6))
plt.hist(all_distances, bins=50, color='skyblue', edgecolor='black')  # More bins for finer granularity
plt.title("Histogram of Distances Between Station Pairs")
plt.xlabel("Distance (minutes)")
plt.ylabel("Number of Journeys")
plt.show()
