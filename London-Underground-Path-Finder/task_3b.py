import openpyxl
from adjacency_list_graph import AdjacencyListGraph
from dijkstra import dijkstra
from print_path import print_path
import re
import matplotlib.pyplot as plt

def normalize_name(name):
    '''Remove extra spaces and convert to lowercase'''
    name = re.sub(r'\s+', ' ', name.strip().lower())
    return name

# load London Underground excel sheet
excel = r'C:\Users\ainho\OneDrive\Escritorio\AADD Project\AA&DE\database.xlsx'
# uses openpyxl to load the excel sheet in our code
wb = openpyxl.load_workbook(excel)
# selects the first sheet from the excel 
sheet = wb.active

# Defining the structures needed as sets to avoid repeated values
lines_set = set()
stations = set()
connections = set()

# iter_rows is a method from openpyxl that allows us to select only the data needed from the excel
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=4):
    # adds the first cell value to the set - column A - line names
    line = row[0].value
    lines_set.add(line)

    station1 = row[1].value  # Column B (first station)
    station2 = row[2].value  # Column C (second station)

    # if column C is empty and column B is not, add the value of column B to the stations set 
    if station2 is None and station1 is not None:
        normalized_station = normalize_name(station1)
        stations.add(normalized_station)

    # if column C is not empty then add a tuple to the connections set
    if station2 is not None:
        normalized_station1 = normalize_name(station1)
        normalized_station2 = normalize_name(station2)
        
        # creates a sorted tuple of stations to ensure unique pairs (A,B and B,A are treated as the same)
        pair = tuple(sorted((normalized_station1, normalized_station2)))
        
        # adds the connection with an unweighted distance of 1, this will help considering only the number of stops 
        # ensuring all distances are treated equaly
        connections.add((pair[0], pair[1], 1))

#* CREATE A GRAPH AND ADD THE EDGES
# convert sets into lists to make data more accesible
connections = list(connections)
stations = list(stations)
# generate a graph with vertices being the stations on the excel
graph = AdjacencyListGraph(len(stations), False, True)
# insert the edges - connections between stations using the provided method from adjacency_list_graph
for edge in connections:
    station1_index = stations.index(edge[0])
    station2_index = stations.index(edge[1])

    # checks if the edge already exists - ensures no repetition
    if not graph.has_edge(station1_index, station2_index):
        graph.insert_edge(station1_index, station2_index, edge[2])


# initialize counter for journeys
journey_counter = 0

# track the longest journey
longest_distance = -1  
longest_path = None
furthest_station_pair = None
all_distances = []  # collect all distances for histogram

# run dijkstra's algorithm for each station
for s1_index in range(len(stations)):
    d, pi = dijkstra(graph, s1_index)
    for s2_index in range(s1_index + 1, len(stations)):
        journey_counter += 1
        if s1_index != s2_index:
            distance = d[s2_index]
            # we add all distances (nº of stops) to 'all_distances' for plotting the histogram later
            all_distances.append(distance)
             # finding the longest distance
            if distance > longest_distance:
                longest_distance = distance
                furthest_station_pair = (stations[s1_index], stations[s2_index])
                longest_path = print_path(pi, s1_index, s2_index, lambda v: stations[v])

# print journey details
print(f"Total number of journeys calculated: {journey_counter}")
print(f"Longest journey duration: {longest_distance} stops")
if furthest_station_pair:
    print(f"Path from {furthest_station_pair[0]} to {furthest_station_pair[1]}: {' -> '.join(longest_path)}")

# Plot histogram of journey stops
plt.figure(figsize=(10, 6))
plt.hist(all_distances, bins=range(1, longest_distance + 2), color='skyblue', edgecolor='black')
plt.title("Histogram of Number of Stops Between Station Pairs")
plt.xlabel("Number of Stops")
plt.ylabel("Number of Journeys")
plt.xticks(range(1, longest_distance + 1))
plt.show()